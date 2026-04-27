"""
╔══════════════════════════════════════════════════════╗
║   ÉLEVAGE OUAKOUBO — Système de Gestion v4.0        ║
║   Streamlit + SQLite + Plotly + Claude AI           ║
╚══════════════════════════════════════════════════════╝
"""

import streamlit as st
import psycopg2
from psycopg2.extras import RealDictCursor
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, datetime, timedelta
import hashlib, io, base64, json, os
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

# ═══════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════
def inject_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');
    :root {
        --bg:#F4F6F9; --surface:#FFFFFF; --border:#E2E8F0;
        --text:#1A202C; --muted:#64748B;
        --sb:#0F172A; --sb2:#1E293B;
        --accent:#16A34A; --acc-l:#DCFCE7; --acc-d:#14532D;
        --warn:#F59E0B; --danger:#EF4444; --info:#3B82F6;
        --r:12px; --sh:0 1px 3px rgba(0,0,0,.06),0 4px 16px rgba(0,0,0,.06);
        --sh-lg:0 8px 30px rgba(0,0,0,.12);
    }
    .stApp{background:var(--bg)!important;font-family:'Sora',sans-serif;}
    .block-container{padding:2rem 2.5rem 3rem!important;max-width:1400px;}
    [data-testid="stSidebar"]{background:var(--sb)!important;border-right:1px solid var(--sb2);}
    [data-testid="stSidebar"]>div{padding-top:0!important;}
    [data-testid="stSidebar"] *{color:#CBD5E1!important;font-family:'Sora',sans-serif;}
    [data-testid="stSidebar"] .stRadio>label{display:none;}
    [data-testid="stSidebar"] .stRadio div[role="radiogroup"]{gap:2px!important;}
    [data-testid="stSidebar"] .stRadio label{
        display:flex!important;align-items:center;gap:10px;
        padding:10px 16px!important;border-radius:8px!important;
        font-size:.88rem!important;font-weight:500!important;
        color:#94A3B8!important;cursor:pointer;transition:all .18s ease;
        border:none!important;background:transparent!important;
    }
    [data-testid="stSidebar"] .stRadio label:hover{background:var(--sb2)!important;color:#E2E8F0!important;}
    h1{font-family:'Sora',sans-serif!important;font-size:1.75rem!important;font-weight:700!important;color:var(--text)!important;letter-spacing:-.5px;}
    h2{font-family:'Sora',sans-serif!important;font-size:1.25rem!important;font-weight:600!important;color:var(--text)!important;margin-top:1.5rem!important;}
    h3{font-size:1rem!important;font-weight:600!important;color:var(--muted)!important;}
    [data-testid="metric-container"]{
        background:var(--surface);border:1px solid var(--border);
        border-radius:var(--r);padding:18px 22px;box-shadow:var(--sh);
        transition:box-shadow .2s,transform .2s;position:relative;overflow:hidden;
    }
    [data-testid="metric-container"]::before{
        content:'';position:absolute;top:0;left:0;right:0;height:3px;
        background:linear-gradient(90deg,var(--accent),#22C55E);
    }
    [data-testid="metric-container"]:hover{box-shadow:var(--sh-lg);transform:translateY(-2px);}
    [data-testid="stMetricValue"]{font-family:'Sora',sans-serif!important;font-size:1.8rem!important;font-weight:700!important;color:var(--text)!important;}
    [data-testid="stMetricLabel"]{font-size:.75rem!important;font-weight:600!important;text-transform:uppercase;letter-spacing:.6px;color:var(--muted)!important;}
    .stButton>button{
        background:var(--accent)!important;color:#fff!important;border:none!important;
        border-radius:8px!important;padding:9px 22px!important;
        font-family:'Sora',sans-serif!important;font-weight:600!important;font-size:.88rem!important;
        transition:all .2s ease!important;box-shadow:0 2px 8px rgba(22,163,74,.25)!important;
    }
    .stButton>button:hover{background:var(--acc-d)!important;transform:translateY(-1px)!important;box-shadow:0 4px 16px rgba(22,163,74,.35)!important;}
    .stTextInput input,.stNumberInput input,.stTextArea textarea,[data-baseweb="select"]>div{
        border-radius:8px!important;border:1.5px solid var(--border)!important;
        font-family:'Sora',sans-serif!important;font-size:.88rem!important;
        background:var(--surface)!important;color:var(--text)!important;
    }
    .stTextInput input:focus,.stNumberInput input:focus,.stTextArea textarea:focus{
        border-color:var(--accent)!important;box-shadow:0 0 0 3px rgba(22,163,74,.12)!important;
    }
    label[data-testid="stWidgetLabel"]{font-size:.82rem!important;font-weight:600!important;color:var(--muted)!important;text-transform:uppercase;letter-spacing:.4px;}
    [data-testid="stForm"]{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);padding:24px;box-shadow:var(--sh);}
    .streamlit-expanderHeader{background:var(--surface)!important;border:1px solid var(--border)!important;border-radius:var(--r)!important;font-family:'Sora',sans-serif!important;font-weight:600!important;font-size:.9rem!important;color:var(--text)!important;padding:14px 18px!important;}
    .streamlit-expanderContent{border:1px solid var(--border)!important;border-top:none!important;border-radius:0 0 var(--r) var(--r)!important;background:var(--surface)!important;}
    .stTabs [data-baseweb="tab-list"]{background:var(--surface)!important;border-radius:var(--r) var(--r) 0 0;border:1px solid var(--border);border-bottom:none;padding:6px 8px 0 8px;gap:4px;}
    .stTabs [data-baseweb="tab"]{font-family:'Sora',sans-serif!important;font-size:.85rem!important;font-weight:600!important;color:var(--muted)!important;padding:8px 16px!important;border-radius:8px 8px 0 0!important;background:transparent!important;}
    .stTabs [aria-selected="true"]{background:var(--bg)!important;color:var(--accent)!important;border-bottom:2px solid var(--accent)!important;}
    .stTabs [data-baseweb="tab-panel"]{background:var(--surface);border:1px solid var(--border);border-radius:0 0 var(--r) var(--r);padding:24px!important;}
    [data-testid="stDataFrame"]{border-radius:var(--r);overflow:hidden;box-shadow:var(--sh);border:1px solid var(--border);}
    [data-testid="stAlert"]{border-radius:8px!important;font-family:'Sora',sans-serif!important;font-size:.88rem!important;border-left-width:4px!important;}
    hr{border:none;border-top:1px solid var(--border);margin:20px 0;}

    /* Page header */
    .ph{display:flex;align-items:center;gap:16px;padding:24px 28px;background:var(--surface);border:1px solid var(--border);border-radius:var(--r);box-shadow:var(--sh);margin-bottom:24px;}
    .ph-icon{width:50px;height:50px;background:var(--acc-l);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.5rem;flex-shrink:0;}
    .ph-title{font-family:'Sora',sans-serif;font-size:1.4rem;font-weight:700;color:var(--text);line-height:1.2;}
    .ph-sub{font-size:.82rem;color:var(--muted);margin-top:2px;}

    /* Badges */
    .badge{display:inline-flex;align-items:center;gap:4px;padding:3px 12px;border-radius:20px;font-size:.75rem;font-weight:600;font-family:'Sora',sans-serif;}
    .bv{background:#DCFCE7;color:#166534;border:1px solid #BBF7D0;}
    .bvd{background:#FEF9C3;color:#854D0E;border:1px solid #FDE68A;}
    .bm{background:#FEE2E2;color:#991B1B;border:1px solid #FECACA;}
    .bml{background:#DBEAFE;color:#1E40AF;border:1px solid #BFDBFE;}
    .bf{background:#FCE7F3;color:#9D174D;border:1px solid #FBCFE8;}
    .bi{background:#EFF6FF;color:#1D4ED8;border:1px solid #BFDBFE;}
    .bw{background:#FEF3C7;color:#92400E;border:1px solid #FDE68A;}

    /* Fiche */
    .fiche{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);overflow:hidden;box-shadow:var(--sh);}
    .fiche-top{background:linear-gradient(135deg,#0F172A,#1E293B);padding:24px 28px;display:flex;align-items:center;gap:16px;}
    .fiche-av{width:60px;height:60px;background:rgba(22,163,74,.2);border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:1.8rem;border:2px solid rgba(22,163,74,.5);flex-shrink:0;}
    .fiche-nom{font-family:'Sora',sans-serif;font-size:1.4rem;font-weight:700;color:#F1F5F9;}
    .fiche-id{font-size:.78rem;color:#64748B;margin-top:2px;font-family:'JetBrains Mono',monospace;}
    .fiche-body{padding:24px 28px;}
    .fiche-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:18px;margin-bottom:20px;}
    .ff label{display:block;font-size:.72rem;font-weight:600;text-transform:uppercase;letter-spacing:.6px;color:var(--muted);margin-bottom:3px;}
    .ff span{font-size:.93rem;font-weight:500;color:var(--text);}

    /* Alert card */
    .alert-card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 18px;margin-bottom:10px;display:flex;align-items:flex-start;gap:12px;box-shadow:var(--sh);}
    .alert-icon{font-size:1.4rem;flex-shrink:0;margin-top:1px;}
    .alert-title{font-weight:600;font-size:.9rem;color:var(--text);}
    .alert-sub{font-size:.8rem;color:var(--muted);margin-top:2px;}
    .alert-danger{border-left:4px solid var(--danger);}
    .alert-warn{border-left:4px solid var(--warn);}
    .alert-info{border-left:4px solid var(--info);}
    .alert-ok{border-left:4px solid var(--accent);}

    /* Chat */
    .chat-user{background:#DCFCE7;border-radius:12px 12px 2px 12px;padding:12px 16px;margin:6px 0 6px 40px;font-size:.9rem;color:#14532D;}
    .chat-ai{background:var(--surface);border:1px solid var(--border);border-radius:12px 12px 12px 2px;padding:12px 16px;margin:6px 40px 6px 0;font-size:.9rem;color:var(--text);box-shadow:var(--sh);}
    .chat-label{font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px;}

    /* ── Masquer tous les widgets radio dans la sidebar ── */
    [data-testid="stSidebar"] .stRadio{display:none!important;}

    /* Sidebar */
    .sb-logo{padding:20px 18px 14px;border-bottom:1px solid #1E293B;margin-bottom:4px;}
    .sb-logo-inner{display:flex;align-items:center;gap:12px;}
    .sb-logo-icon{width:38px;height:38px;background:rgba(22,163,74,.15);border:1px solid rgba(22,163,74,.3);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:1.3rem;}
    .sb-logo-name{font-family:'Sora',sans-serif;font-size:.92rem;font-weight:700;color:#E2E8F0;line-height:1.2;}
    .sb-logo-sub{font-size:.68rem;color:#475569;}

    /* Accordion section header */
    .sb-acc-header{
        display:flex;align-items:center;justify-content:space-between;
        padding:10px 18px 8px;
        cursor:pointer;
        font-size:.63rem;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;
        color:#475569;
        border-top:1px solid #1E293B;
        margin-top:4px;
        transition:color .15s;
        user-select:none;
    }
    .sb-acc-header:hover{color:#94A3B8;}
    .sb-acc-header.open{color:#74C69D;}
    .sb-acc-arrow{font-size:.7rem;transition:transform .2s;}
    .sb-acc-header.open .sb-acc-arrow{transform:rotate(90deg);}

    /* Nav item bar */
    .sb-nav-item{
        display:flex;align-items:center;gap:10px;
        padding:9px 18px 9px 26px;
        font-size:.85rem;font-weight:500;
        color:#94A3B8;
        cursor:pointer;
        border-left:3px solid transparent;
        transition:all .15s ease;
        text-decoration:none;
        white-space:nowrap;
    }
    .sb-nav-item:hover{background:#1E293B;color:#E2E8F0;border-left-color:#334155;}
    .sb-nav-item.active{
        background:rgba(22,163,74,.1);
        color:#4ADE80;
        border-left-color:#16A34A;
        font-weight:600;
    }
    .sb-nav-icon{font-size:1rem;flex-shrink:0;width:18px;text-align:center;}
    .sb-nav-badge{
        margin-left:auto;background:#EF4444;color:white;
        border-radius:10px;padding:1px 7px;font-size:.65rem;font-weight:700;
    }

    /* User footer */
    .sb-user{padding:10px 18px;display:flex;align-items:center;gap:10px;border-top:1px solid #1E293B;margin-top:auto;}
    .sb-user-av{width:28px;height:28px;background:rgba(22,163,74,.2);border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:.85rem;}
    .sb-user-name{font-size:.78rem;font-weight:600;color:#94A3B8;}
    .sb-user-role{font-size:.65rem;color:#475569;}

    /* Déconnexion btn sidebar */
    [data-testid="stSidebar"] .stButton>button{
        background:#1E293B!important;
        color:#94A3B8!important;
        border:1px solid #334155!important;
        font-size:.8rem!important;
        padding:7px 14px!important;
        box-shadow:none!important;
        margin:6px 18px 10px!important;
        width:calc(100% - 36px)!important;
    }
    [data-testid="stSidebar"] .stButton>button:hover{
        background:#EF4444!important;color:#fff!important;border-color:#EF4444!important;
    }

    /* Login */
    .lw{max-width:420px;margin:50px auto;background:var(--surface);border:1px solid var(--border);border-radius:16px;box-shadow:var(--sh-lg);overflow:hidden;}
    .lt{background:linear-gradient(135deg,#0F172A,#1E293B);padding:36px 36px 28px;text-align:center;}
    .ll{width:68px;height:68px;background:rgba(22,163,74,.15);border:2px solid rgba(22,163,74,.4);border-radius:16px;display:flex;align-items:center;justify-content:center;font-size:2rem;margin:0 auto 14px;}
    .ltitle{font-family:'Sora',sans-serif;font-size:1.4rem;font-weight:700;color:#F1F5F9;}
    .lsub{font-size:.8rem;color:#64748B;margin-top:3px;}
    .lb{padding:28px 36px 32px;}

    /* Genealogy tree */
    .tree-node{background:var(--surface);border:1.5px solid var(--border);border-radius:10px;padding:10px 16px;text-align:center;min-width:120px;box-shadow:var(--sh);}
    .tree-node-main{border-color:var(--accent);border-width:2px;}
    </style>
    """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════
# BASE DE DONNÉES — PostgreSQL (Supabase)
# ═══════════════════════════════════════════════════════

@st.cache_resource
def get_conn():
    """Connexion persistante à Supabase PostgreSQL via st.secrets."""
    return psycopg2.connect(
        st.secrets["DATABASE_URL"],
        cursor_factory=RealDictCursor,
        sslmode="require"
    )

def run_query(sql, params=None, fetch=False):
    """Exécute une requête SQL. Reconnecte automatiquement si besoin."""
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            if fetch:
                return cur.fetchall()
            # committed via run_query
    except Exception:
        conn.rollback()
        # Reset cache et réessai
        st.cache_resource.clear()
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            if fetch:
                return cur.fetchall()
            # committed via run_query

def read_sql(sql, params=None):
    """Équivalent de pd.read_sql pour PostgreSQL."""
    rows = run_query(sql, params, fetch=True)
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame([dict(r) for r in rows])

def init_db():
    """Crée les tables si elles n'existent pas (PostgreSQL syntax)."""
    conn = get_conn()
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS utilisateur (
                id SERIAL PRIMARY KEY,
                login TEXT UNIQUE,
                password TEXT,
                role TEXT DEFAULT 'admin',
                datecreation DATE DEFAULT CURRENT_DATE
            );
            CREATE TABLE IF NOT EXISTS lapin (
                idlapin SERIAL PRIMARY KEY,
                nom TEXT, sexe TEXT, race TEXT, couleur TEXT,
                datenaissance DATE, origine TEXT,
                idpere INTEGER, idmere INTEGER,
                statut TEXT DEFAULT 'vivant',
                remarques TEXT, photo TEXT
            );
            CREATE TABLE IF NOT EXISTS reproduction (
                idrepro SERIAL PRIMARY KEY,
                idmale INTEGER, idfemelle INTEGER,
                dateaccouplement DATE, datecontrole DATE,
                datemisebas DATE,
                gestationconfirmee INTEGER DEFAULT 0, notes TEXT
            );
            CREATE TABLE IF NOT EXISTS portee (
                idportee SERIAL PRIMARY KEY,
                idrepro INTEGER, datenaissance DATE,
                nbvivant INTEGER, nbmort INTEGER, notes TEXT
            );
            CREATE TABLE IF NOT EXISTS pesee (
                idpesee SERIAL PRIMARY KEY,
                idlapin INTEGER, datepesee DATE, poids REAL
            );
            CREATE TABLE IF NOT EXISTS sante (
                idsante SERIAL PRIMARY KEY,
                idlapin INTEGER, datetraitement DATE,
                typetraitement TEXT, produit TEXT, remarque TEXT,
                daterappel DATE
            );
            CREATE TABLE IF NOT EXISTS vente (
                idvente SERIAL PRIMARY KEY,
                idlapin INTEGER, datevente DATE, prix REAL, client TEXT
            );
            CREATE TABLE IF NOT EXISTS stock (
                idstock SERIAL PRIMARY KEY,
                type TEXT, nom TEXT, quantite REAL, unite TEXT,
                prixunitaire REAL, dateachat DATE, notes TEXT
            );
            CREATE TABLE IF NOT EXISTS journal (
                idjournal SERIAL PRIMARY KEY,
                dateaction TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                utilisateur TEXT, action TEXT, detail TEXT
            );
        """)
        # Admin par défaut
        pwd = hashlib.sha256("ouakoubo2025".encode()).hexdigest()
        cur.execute("""
            INSERT INTO utilisateur (login, password, role)
            VALUES (%s, %s, %s)
            ON CONFLICT (login) DO NOTHING
        """, ("admin", pwd, "admin"))
    # committed via run_query

def log_action(action, detail=""):
    run_query(
        "INSERT INTO journal (utilisateur, action, detail) VALUES (%s, %s, %s)",
        (st.session_state.get("username","%s"), action, detail)
    )

# ═══════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════
PLOTLY = dict(
    paper_bgcolor='white', plot_bgcolor='white',
    font_family='Sora', font_color='#1A202C',
    margin=dict(t=46,b=22,l=14,r=14),
    title_font_size=14, title_font_color='#1A202C',
    legend=dict(bgcolor='rgba(0,0,0,0)', font_size=11),
)
GREENS = ['#14532D','#166534','#16A34A','#22C55E','#4ADE80','#86EFAC']

def pchart(fig):
    fig.update_layout(**PLOTLY)
    st.plotly_chart(fig, use_container_width=True, config={'displayModeBar':False})

def ph(icon, title, sub):
    st.markdown(f'<div class="ph"><div class="ph-icon">{icon}</div><div><div class="ph-title">{title}</div><div class="ph-sub">{sub}</div></div></div>', unsafe_allow_html=True)

def badge(txt, cls):
    return f'<span class="badge {cls}">{txt}</span>'

def statut_badge(s):
    m = {'vivant':'bv','vendu':'bvd','mort':'bm'}
    return badge(s.upper(), m.get(s,'bi'))

def role_ok(required):
    role = st.session_state.get("role","")
    if required == "admin" and role != "admin":
        st.warning("⛔ Accès réservé aux administrateurs.")
        return False
    if required == "employe" and role not in ["admin","employe"]:
        st.warning("⛔ Accès non autorisé.")
        return False
    return True

# ═══════════════════════════════════════════════════════
# EXPORT EXCEL
# ═══════════════════════════════════════════════════════
def export_excel(dfs: dict, title="Export") -> bytes:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    green = "16A34A"; dark = "0F172A"; light = "DCFCE7"
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_fill  = PatternFill("solid", fgColor=dark)
    sub_fill  = PatternFill("solid", fgColor=green)
    alt_fill  = PatternFill("solid", fgColor="F4F6F9")
    thin      = Border(left=Side(style='thin',color='E2E8F0'),right=Side(style='thin',color='E2E8F0'),
                       top=Side(style='thin',color='E2E8F0'),bottom=Side(style='thin',color='E2E8F0'))
    for sheet_name, df in dfs.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns),1))
        tc = ws.cell(1,1, value=f"Élevage Ouakoubo — {sheet_name}")
        tc.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
        tc.fill = PatternFill("solid", fgColor=dark)
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 26
        # Date row
        ws.cell(2,1, value=f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
        ws.cell(2,1).font = Font(italic=True, color="64748B", name="Calibri", size=9)
        # Headers
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(3, ci, value=str(col))
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=green)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin
        ws.row_dimensions[3].height = 20
        # Data
        for ri, row in enumerate(df.itertuples(index=False), 4):
            fill = alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, value=val)
                c.font = Font(name="Calibri", size=10)
                c.fill = fill
                c.alignment = Alignment(vertical='center')
                c.border = thin
        # Auto width
        for ci, col in enumerate(df.columns, 1):
            max_len = max([len(str(col))] + [len(str(v)) for v in df[col].values])
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)
    wb.save(buf)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════
# EXPORT PDF — Fiche lapin
# ═══════════════════════════════════════════════════════
def export_pdf_fiche(row, pesees_df, sante_df) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    GREEN = colors.HexColor("#16A34A")
    DARK  = colors.HexColor("#0F172A")
    GREY  = colors.HexColor("#64748B")
    LIGHT = colors.HexColor("#F4F6F9")

    title_style = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=20, textColor=DARK, spaceAfter=4)
    sub_style   = ParagraphStyle('s', fontName='Helvetica', fontSize=10, textColor=GREY, spaceAfter=16)
    label_style = ParagraphStyle('l', fontName='Helvetica-Bold', fontSize=8, textColor=GREY, spaceAfter=2)
    val_style   = ParagraphStyle('v', fontName='Helvetica', fontSize=11, textColor=DARK, spaceAfter=12)
    sec_style   = ParagraphStyle('sec', fontName='Helvetica-Bold', fontSize=12, textColor=GREEN, spaceBefore=16, spaceAfter=8)

    elems = []
    # Header
    elems.append(Paragraph(f"🐇 Fiche — {row['nom']}", title_style))
    elems.append(Paragraph(f"ID #{row['idlapin']} · Statut : {row['statut'].upper()} · Généré le {datetime.now().strftime('%d/%m/%Y')}", sub_style))
    elems.append(HRFlowable(width="100%", thickness=2, color=GREEN, spaceAfter=14))

    # Info grid
    infos = [
        ("Sexe", "Mâle ♂️" if row['sexe']=='M' else "Femelle ♀️"),
        ("Race", row['race'] or "—"),
        ("Couleur", row['couleur'] or "—"),
        ("Date de naissance", str(row['datenaissance'] or "—")),
        ("Origine", row['origine'] or "—"),
        ("Remarques", row['remarques'] or "—"),
    ]
    tdata = [[Paragraph(f"<b>{k}</b>", ParagraphStyle('k',fontName='Helvetica-Bold',fontSize=9,textColor=GREY)),
              Paragraph(str(v), ParagraphStyle('val',fontName='Helvetica',fontSize=10,textColor=DARK))]
             for k,v in infos]
    t = Table(tdata, colWidths=[5*cm, 12*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),LIGHT),
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[colors.white, LIGHT]),
        ('GRID',(0,0),(-1,-1),.4,colors.HexColor("#E2E8F0")),
        ('PADDING',(0,0),(-1,-1),8),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    elems.append(t)

    # Pesées
    if not pesees_df.empty:
        elems.append(Paragraph("Historique des pesées", sec_style))
        ph_data = [["Date", "Poids (g)"]] + [[str(r['datepesee']), f"{r['poids']:.0f}g"] for _,r in pesees_df.iterrows()]
        pt = Table(ph_data, colWidths=[8*cm,8*cm])
        pt.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),GREEN),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,0),10),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, LIGHT]),
            ('GRID',(0,0),(-1,-1),.4,colors.HexColor("#E2E8F0")),
            ('PADDING',(0,0),(-1,-1),7),
            ('ALIGN',(1,0),(-1,-1),'CENTER'),
        ]))
        elems.append(pt)

    # Santé
    if not sante_df.empty:
        elems.append(Paragraph("Historique santé", sec_style))
        sh_data = [["Date","Type","Produit"]] + [[str(r['datetraitement']),r['typetraitement'],r['produit'] or '—'] for _,r in sante_df.iterrows()]
        st2 = Table(sh_data, colWidths=[5*cm,5*cm,7*cm])
        st2.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),GREEN),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, LIGHT]),
            ('GRID',(0,0),(-1,-1),.4,colors.HexColor("#E2E8F0")),
            ('PADDING',(0,0),(-1,-1),7),
        ]))
        elems.append(st2)

    # Footer
    elems.append(Spacer(1, 0.5*cm))
    elems.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
    elems.append(Paragraph("Élevage Ouakoubo — Système de gestion cuniculicole v4.0", ParagraphStyle('f',fontName='Helvetica',fontSize=8,textColor=GREY,spaceAfter=0)))
    doc.build(elems)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════
# RAPPORT MENSUEL PDF
# ═══════════════════════════════════════════════════════
def export_rapport_mensuel(mois_str: str) -> bytes:
    conn = get_conn()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    GREEN = colors.HexColor("#16A34A"); DARK = colors.HexColor("#0F172A"); GREY = colors.HexColor("#64748B"); LIGHT = colors.HexColor("#F4F6F9")
    sec_style = ParagraphStyle('sec', fontName='Helvetica-Bold', fontSize=13, textColor=GREEN, spaceBefore=18, spaceAfter=8)
    title_style = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=22, textColor=DARK, spaceAfter=4)
    sub_style = ParagraphStyle('s', fontName='Helvetica', fontSize=10, textColor=GREY, spaceAfter=14)

    total = run_query("SELECT COUNT(*) FROM lapin WHERE statut='vivant'").fetchone()[0]
    portees_mois = run_query("SELECT COUNT(*), COALESCE(SUM(nbvivant),0), COALESCE(SUM(nbmort),0) FROM portee WHERE TO_CHAR(datenaissance, 'YYYY-MM')=%s", (mois_str,)).fetchone()
    ventes_mois = run_query("SELECT COUNT(*), COALESCE(SUM(prix),0) FROM vente WHERE TO_CHAR(datevente, 'YYYY-MM')=%s", (mois_str,)).fetchone()
    traitements = run_query("SELECT COUNT(*) FROM sante WHERE TO_CHAR(datetraitement, 'YYYY-MM')=%s", (mois_str,)).fetchone()[0]

    elems = []
    elems.append(Paragraph(f"Rapport Mensuel — {mois_str}", title_style))
    elems.append(Paragraph(f"Élevage Ouakoubo · Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", sub_style))
    elems.append(HRFlowable(width="100%", thickness=2, color=GREEN, spaceAfter=14))

    kpis = [
        ["Indicateur", "Valeur"],
        ["Total lapins vivants", str(total)],
        ["Portées ce mois", str(portees_mois[0])],
        ["Lapereaux nés vivants", str(portees_mois[1])],
        ["Mort-nés", str(portees_mois[2])],
        ["Ventes réalisées", str(ventes_mois[0])],
        ["Revenus (FCFA)", f"{ventes_mois[1]:,.0f}"],
        ["Traitements effectués", str(traitements)],
    ]
    kt = Table(kpis, colWidths=[10*cm, 7*cm])
    kt.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),DARK),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),10),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, LIGHT]),
        ('GRID',(0,0),(-1,-1),.4,colors.HexColor("#E2E8F0")),
        ('PADDING',(0,0),(-1,-1),9),
        ('ALIGN',(1,0),(-1,-1),'CENTER'),
    ]))
    elems.append(Paragraph("Indicateurs clés", sec_style))
    elems.append(kt)

    # Ventes détail
    df_v = read_sql("SELECT l.nom, v.datevente, v.prix, v.client FROM vente v JOIN lapin l ON v.idlapin=l.idlapin WHERE TO_CHAR(v.datevente, 'YYYY-MM')=%s", conn, params=[mois_str])
    if not df_v.empty:
        elems.append(Paragraph("Détail des ventes", sec_style))
        vd = [["Lapin","Date","Prix (FCFA)","Client"]] + [[r['nom'],str(r['datevente']),f"{r['prix']:,.0f}",r['client'] or '—'] for _,r in df_v.iterrows()]
        vt = Table(vd, colWidths=[4*cm,4*cm,4*cm,5*cm])
        vt.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),GREEN),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, LIGHT]),
            ('GRID',(0,0),(-1,-1),.4,colors.HexColor("#E2E8F0")),
            ('PADDING',(0,0),(-1,-1),7),
        ]))
        elems.append(vt)

    elems.append(Spacer(1,0.5*cm))
    elems.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor("#E2E8F0")))
    elems.append(Paragraph("Élevage Ouakoubo — Système de gestion cuniculicole v4.0", ParagraphStyle('f',fontName='Helvetica',fontSize=8,textColor=GREY)))
    doc.build(elems)
    # connection pooled
    return buf.getvalue()

# ═══════════════════════════════════════════════════════
# ALERTES
# ═══════════════════════════════════════════════════════
def get_alertes():
    conn = get_conn()
    alertes = []
    today = date.today()

    # Gestation en attente (J+28 = mise bas probable)
    repros = read_sql("""
        SELECT r.idrepro, lm.nom AS male, lf.nom AS femelle,
               r.dateaccouplement, r.gestationconfirmee
        FROM reproduction r
        JOIN lapin lm ON r.idmale=lm.idlapin
        JOIN lapin lf ON r.idfemelle=lf.idlapin
        WHERE r.gestationconfirmee=1
    """)
    for _, r in repros.iterrows():
        d = datetime.strptime(str(r['dateaccouplement']), '%Y-%m-%d').date()
        j28 = d + timedelta(days=28)
        diff = (j28 - today).days
        if -2 <= diff <= 5:
            alertes.append({"type":"danger", "icon":"🍼",
                "titre": f"Mise bas imminente — {r['femelle']}",
                "detail": f"Accouplement {r['male']} × {r['femelle']} le {d.strftime('%d/%m/%Y')} · Mise bas prévue le {j28.strftime('%d/%m/%Y')}"})
        elif diff <= 0 and diff < -2:
            alertes.append({"type":"warn", "icon":"⚠️",
                "titre": f"Portée en retard — {r['femelle']}",
                "detail": f"Mise bas attendue le {j28.strftime('%d/%m/%Y')} · {abs(diff)} jour(s) de retard"})

    # Contrôle gestation à faire
    ctrl = read_sql("""
        SELECT r.datecontrole, lm.nom AS male, lf.nom AS femelle
        FROM reproduction r
        JOIN lapin lm ON r.idmale=lm.idlapin
        JOIN lapin lf ON r.idfemelle=lf.idlapin
        WHERE r.gestationconfirmee=0 AND r.datecontrole IS NOT NULL
    """)
    for _, r in ctrl.iterrows():
        try:
            dc = datetime.strptime(str(r['datecontrole']), '%Y-%m-%d').date()
            diff = (dc - today).days
            if -3 <= diff <= 3:
                alertes.append({"type":"warn", "icon":"🔬",
                    "titre": f"Contrôle gestation — {r['femelle']}",
                    "detail": f"Prévu le {dc.strftime('%d/%m/%Y')} pour {r['male']} × {r['femelle']}"})
        except: pass

    # Rappels vaccin/traitement
    rappels = read_sql("""
        SELECT l.nom, s.typetraitement, s.daterappel
        FROM sante s JOIN lapin l ON s.idlapin=l.idlapin
        WHERE s.daterappel IS NOT NULL
    """)
    for _, r in rappels.iterrows():
        try:
            dr = datetime.strptime(str(r['daterappel']), '%Y-%m-%d').date()
            diff = (dr - today).days
            if 0 <= diff <= 7:
                alertes.append({"type":"warn","icon":"💉",
                    "titre": f"Rappel {r['typetraitement']} — {r['nom']}",
                    "detail": f"Prévu le {dr.strftime('%d/%m/%Y')} (dans {diff} jour(s))"})
            elif diff < 0:
                alertes.append({"type":"danger","icon":"🚨",
                    "titre": f"Rappel EN RETARD — {r['nom']}",
                    "detail": f"{r['typetraitement']} prévu le {dr.strftime('%d/%m/%Y')} · {abs(diff)} jour(s) de retard"})
        except: pass

    # Lapins sans pesée depuis 14 jours
    lapins = read_sql("SELECT idlapin, nom FROM lapin WHERE statut='vivant'")
    for _, l in lapins.iterrows():
        last = run_query("SELECT MAX(datepesee) FROM pesee WHERE idlapin=%s", (int(l['idlapin']),)).fetchone()[0]
        if last:
            dl = datetime.strptime(last, '%Y-%m-%d').date()
            if (today - dl).days > 14:
                alertes.append({"type":"info","icon":"⚖️",
                    "titre": f"Pesée requise — {l['nom']}",
                    "detail": f"Dernière pesée il y a {(today-dl).days} jours"})

    # connection pooled
    return alertes

# ═══════════════════════════════════════════════════════
# ASSISTANT IA
# ═══════════════════════════════════════════════════════
def call_claude(messages, system_prompt):
    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json"},
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 1000,
                "system": system_prompt,
                "messages": messages
            },
            timeout=30
        )
        data = resp.json()
        return data["content"][0]["text"]
    except Exception as e:
        return f"Erreur de connexion à l'IA : {str(e)}"

def get_elevage_context():
    conn = get_conn()
    ctx = {}
    ctx['total_vivants'] = run_query("SELECT COUNT(*) FROM lapin WHERE statut='vivant'").fetchone()[0]
    ctx['males'] = run_query("SELECT COUNT(*) FROM lapin WHERE sexe='M' AND statut='vivant'").fetchone()[0]
    ctx['femelles'] = run_query("SELECT COUNT(*) FROM lapin WHERE sexe='F' AND statut='vivant'").fetchone()[0]
    ctx['total_portees'] = run_query("SELECT COUNT(*) FROM portee").fetchone()[0]
    ctx['revenus_total'] = run_query("SELECT COALESCE(SUM(prix),0) FROM vente").fetchone()[0]
    ctx['races'] = [r[0] for r in run_query("SELECT DISTINCT race FROM lapin WHERE race != '' AND race IS NOT NULL").fetchall()]

    # Lapins liste
    lapins = run_query("SELECT nom, sexe, race, statut, datenaissance FROM lapin WHERE statut='vivant' LIMIT 20").fetchall()
    ctx['lapins'] = [{"nom":l[0],"sexe":l[1],"race":l[2],"statut":l[3],"naissance":l[4]} for l in lapins]

    # Meilleure femelle (plus de vivants)
    best = run_query("""
        SELECT l.nom, SUM(p.nbvivant) as total
        FROM portee p JOIN reproduction r ON p.idrepro=r.idrepro
        JOIN lapin l ON r.idfemelle=l.idlapin
        GROUP BY l.idlapin ORDER BY total DESC LIMIT 1
    """).fetchone()
    ctx['meilleure_femelle'] = {"nom": best[0], "vivants": best[1]} if best else None

    # connection pooled
    return ctx

# ═══════════════════════════════════════════════════════
# LOGIN
# ═══════════════════════════════════════════════════════
def login_page():
    st.markdown('<div class="lw"><div class="lt"><div class="ll">🐇</div><div class="ltitle">Élevage Ouakoubo</div><div class="lsub">Système de gestion cuniculicole v4.0</div></div><div class="lb">', unsafe_allow_html=True)
    _, col, _ = st.columns([1,3,1])
    with col:
        with st.form("lf"):
            login    = st.text_input("Identifiant")
            password = st.text_input("Mot de passe", type="password")
            ok = st.form_submit_button("🔓  Connexion", use_container_width=True)
        if ok:
            h = hashlib.sha256(password.encode()).hexdigest()
            conn = get_conn()
            user = run_query("SELECT login,role FROM utilisateur WHERE login=%s AND password=%s", (login,h)).fetchone()
            # connection pooled
            if user:
                st.session_state.update({"logged_in":True,"username":user[0],"role":user[1]})
                log_action("CONNEXION", f"Utilisateur {login} connecté")
                st.rerun()
            else:
                st.error("❌ Identifiant ou mot de passe incorrect")
        st.caption("🔑 Défaut : **admin** / **ouakoubo2025**")
    st.markdown("</div></div>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════
def page_dashboard():
    ph("📊", "Tableau de bord", "Vue d'ensemble en temps réel de votre élevage")
    conn = get_conn()

    total    = run_query("SELECT COUNT(*) FROM lapin WHERE statut='vivant'").fetchone()[0]
    males    = run_query("SELECT COUNT(*) FROM lapin WHERE sexe='M' AND statut='vivant'").fetchone()[0]
    femelles = run_query("SELECT COUNT(*) FROM lapin WHERE sexe='F' AND statut='vivant'").fetchone()[0]
    portees  = run_query("SELECT COUNT(*) FROM portee").fetchone()[0]
    ventes   = run_query("SELECT COALESCE(SUM(prix),0) FROM vente").fetchone()[0]
    traite   = run_query("SELECT COUNT(*) FROM sante").fetchone()[0]
    nb_vivants_total = run_query("SELECT COALESCE(SUM(nbvivant),0) FROM portee").fetchone()[0]
    taux_mort = round((run_query("SELECT COALESCE(SUM(nbmort),0) FROM portee").fetchone()[0] / max(nb_vivants_total+1,1))*100, 1) if nb_vivants_total else 0

    c1,c2,c3,c4,c5,c6,c7,c8 = st.columns(8)
    c1.metric("🐇 Vivants", total)
    c2.metric("♂️ Mâles", males)
    c3.metric("♀️ Femelles", femelles)
    c4.metric("👶 Portées", portees)
    c5.metric("🐣 Lapereaux", nb_vivants_total)
    c6.metric("💰 Revenus", f"{ventes:,.0f}")
    c7.metric("💊 Soins", traite)
    c8.metric("💀 Mortalité", f"{taux_mort}%")

    # Alertes rapides
    alertes = get_alertes()
    if alertes:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown(f"### 🔔 Alertes actives ({len(alertes)})")
        cols = st.columns(min(len(alertes), 3))
        for i, a in enumerate(alertes[:3]):
            cls = {"danger":"alert-danger","warn":"alert-warn","info":"alert-info"}.get(a['type'],"alert-ok")
            cols[i].markdown(f'<div class="alert-card {cls}"><div class="alert-icon">{a["icon"]}</div><div><div class="alert-title">{a["titre"]}</div><div class="alert-sub">{a["detail"]}</div></div></div>', unsafe_allow_html=True)
        if len(alertes) > 3:
            st.caption(f"+ {len(alertes)-3} autre(s) alerte(s) — voir section Alertes")

    st.markdown("<hr>", unsafe_allow_html=True)

    # KPIs avancés
    moy_portee = round(run_query("SELECT AVG(nbvivant) FROM portee").fetchone()[0] or 0, 1)
    best_f = run_query("""SELECT l.nom, SUM(p.nbvivant) as t FROM portee p
        JOIN reproduction r ON p.idrepro=r.idrepro JOIN lapin l ON r.idfemelle=l.idlapin
        GROUP BY l.idlapin ORDER BY t DESC LIMIT 1""").fetchone()
    dernier_revenu = run_query("SELECT COALESCE(SUM(prix),0) FROM vente WHERE TO_CHAR(datevente, 'YYYY-MM')=TO_CHAR('now', 'YYYY-MM')").fetchone()[0]
    c9,c10,c11 = st.columns(3)
    c9.metric("📊 Prolificité moy.", f"{moy_portee} lap./portée")
    c10.metric("🏆 Meilleure reproductrice", best_f[0] if best_f else "—")
    c11.metric("💵 Revenus ce mois", f"{dernier_revenu:,.0f} FCFA")

    st.markdown("<hr>", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        df_r = read_sql("SELECT race, COUNT(*) as n FROM lapin WHERE statut='vivant' AND race!='' GROUP BY race", conn)
        if not df_r.empty:
            fig = px.pie(df_r, values='n', names='race', title="Répartition par race", color_discrete_sequence=GREENS, hole=.4)
            fig.update_traces(textposition='inside', textinfo='percent+label', marker=dict(line=dict(color='white',width=2)))
            pchart(fig)
        else:
            st.info("Ajoutez des lapins avec une race.")
    with col2:
        df_sx = pd.DataFrame({'Sexe':['♂️ Mâles','♀️ Femelles'],'N':[males,femelles]})
        f2 = px.bar(df_sx, x='Sexe', y='N', title="Mâles vs Femelles", color='Sexe',
                    color_discrete_sequence=['#1E40AF','#9D174D'], text='N')
        f2.update_traces(textposition='outside', marker_line_width=0)
        f2.update_layout(showlegend=False, xaxis_title='', yaxis_title='')
        pchart(f2)

    col3, col4 = st.columns(2)
    with col3:
        df_p = read_sql("SELECT datepesee, AVG(poids) as p FROM pesee GROUP BY datepesee ORDER BY datepesee", conn)
        if not df_p.empty:
            f3 = px.area(df_p, x='datepesee', y='p', title="Poids moyen (g)", color_discrete_sequence=['#16A34A'])
            f3.update_traces(fill='tozeroy', fillcolor='rgba(22,163,74,.1)', line_width=2)
            f3.update_layout(xaxis_title='Date', yaxis_title='Poids (g)')
            pchart(f3)
        else:
            st.info("Aucune pesée enregistrée.")
    with col4:
        df_v = read_sql("SELECT TO_CHAR(datevente, 'YYYY-MM') as m, SUM(prix) as t FROM vente GROUP BY m ORDER BY m", conn)
        if not df_v.empty:
            f4 = px.bar(df_v, x='m', y='t', title="Revenus mensuels (FCFA)", color_discrete_sequence=['#16A34A'], text='t')
            f4.update_traces(texttemplate='%{text:,.0f}', textposition='outside', marker_line_width=0)
            f4.update_layout(xaxis_title='', yaxis_title='FCFA')
            pchart(f4)
        else:
            st.info("Aucune vente enregistrée.")

    df_port = read_sql("SELECT TO_CHAR(datenaissance, 'YYYY-MM') as m, COUNT(*) as p, SUM(nbvivant) as v, SUM(nbmort) as mo FROM portee GROUP BY m ORDER BY m", conn)
    if not df_port.empty:
        f5 = go.Figure()
        f5.add_trace(go.Bar(x=df_port['m'], y=df_port['v'], name='Vivants', marker_color='#16A34A', marker_line_width=0))
        f5.add_trace(go.Bar(x=df_port['m'], y=df_port['mo'], name='Mort-nés', marker_color='#EF4444', marker_line_width=0))
        f5.update_layout(barmode='stack', title='Naissances par mois', xaxis_title='', yaxis_title='Lapereaux', **PLOTLY)
        st.plotly_chart(f5, use_container_width=True, config={'displayModeBar':False})

    # Export dashboard
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown("**📥 Exporter les données**")
    c_ex1, c_ex2, _ = st.columns([1,1,3])
    with c_ex1:
        df_all = read_sql("SELECT * FROM lapin")
        xls = export_excel({"Lapins":df_all}, "Tous les lapins")
        st.download_button("📊 Excel — Lapins", xls, "lapins_ouakoubo.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c_ex2:
        mois_dispo = read_sql("SELECT DISTINCT TO_CHAR(datevente, 'YYYY-MM') as m FROM vente WHERE datevente IS NOT NULL ORDER BY m DESC", conn)
        if not mois_dispo.empty:
            mois_sel = c_ex2.selectbox("Mois du rapport PDF", mois_dispo['m'].tolist(), label_visibility="collapsed")
            pdf_b = export_rapport_mensuel(mois_sel)
            st.download_button("📄 Rapport PDF", pdf_b, f"rapport_{mois_sel}.pdf", "application/pdf")
    # connection pooled

# ═══════════════════════════════════════════════════════
# ALERTES
# ═══════════════════════════════════════════════════════
def page_alertes():
    ph("🔔", "Alertes & Rappels", "Surveillance automatique de votre élevage")
    alertes = get_alertes()
    if not alertes:
        st.success("✅ Aucune alerte en cours. Tout va bien !")
        return
    type_colors = {"danger":"alert-danger","warn":"alert-warn","info":"alert-info","ok":"alert-ok"}
    for a in alertes:
        cls = type_colors.get(a['type'],'alert-ok')
        st.markdown(f'<div class="alert-card {cls}"><div class="alert-icon">{a["icon"]}</div><div><div class="alert-title">{a["titre"]}</div><div class="alert-sub">{a["detail"]}</div></div></div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════
# LAPINS
# ═══════════════════════════════════════════════════════
def page_lapins():
    ph("🐇", "Gestion des Lapins", "Registre complet de votre cheptel")
    conn = get_conn()
    tabs = st.tabs(["📋 Liste","➕ Ajouter","🔍 Fiche","🌳 Généalogie"])

    with tabs[0]:
        cf1,cf2,cf3 = st.columns(3)
        fs = cf1.selectbox("Sexe", ["Tous","M","F"])
        ft = cf2.selectbox("Statut", ["Tous","vivant","vendu","mort"])
        fr = cf3.text_input("Recherche race / nom", placeholder="Tapez ici...")
        q="SELECT * FROM lapin WHERE 1=1"; p=[]
        if fs!="Tous": q+=" AND sexe=%s"; p.append(fs)
        if ft!="Tous": q+=" AND statut=%s"; p.append(ft)
        if fr: q+=" AND (race LIKE %s OR nom LIKE %s)"; p.extend([f"%{fr}%",f"%{fr}%"])
        df = read_sql(q, p)
        st.markdown(f'<span class="badge bi">{len(df)} lapin(s)</span>', unsafe_allow_html=True)
        st.dataframe(df.drop(columns=['photo'], errors='ignore'), use_container_width=True, height=360)
        xls = export_excel({"Lapins":df.drop(columns=['photo'],errors='ignore')})
        st.download_button("📊 Exporter Excel", xls, "lapins_filtrés.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tabs[1]:
        if not role_ok("employe"): pass
        else:
            with st.form("fadd"):
                c1,c2,c3=st.columns(3)
                nom=c1.text_input("Nom *"); sexe=c2.selectbox("Sexe *",["M","F"]); race=c3.text_input("Race")
                couleur=c1.text_input("Couleur"); dob=c2.date_input("Date de naissance",value=date.today())
                origine=c3.selectbox("Origine",["Achat","Naissance interne","Don","Autre"])
                all_l=read_sql("SELECT idlapin,nom,sexe FROM lapin")
                c4,c5=st.columns(2)
                ps=c4.selectbox("Père",["—"]+all_l[all_l.sexe=='M']['nom'].tolist())
                ms=c5.selectbox("Mère",["—"]+all_l[all_l.sexe=='F']['nom'].tolist())
                rem=st.text_area("Remarques")
                photo=st.file_uploader("📸 Photo (optionnel)",type=['jpg','jpeg','png'])
                if st.form_submit_button("💾 Enregistrer",use_container_width=True):
                    if not nom: st.error("Le nom est obligatoire.")
                    else:
                        id_p=int(all_l[all_l.nom==ps]['idlapin'].values[0]) if ps!="—" else None
                        id_m=int(all_l[all_l.nom==ms]['idlapin'].values[0]) if ms!="—" else None
                        ph64=base64.b64encode(photo.read()).decode() if photo else None
                        run_query("INSERT INTO lapin (nom,sexe,race,couleur,datenaissance,origine,idpere,idmere,remarques,photo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                     (nom,sexe,race,couleur,str(dob),origine,id_p,id_m,rem,ph64))
                        # committed via run_query
                        log_action("AJOUT_LAPIN", f"Lapin {nom} ajouté")
                        st.success(f"✅ **{nom}** ajouté !"); st.balloons()

    with tabs[2]:
        all_l2=read_sql("SELECT idlapin,nom FROM lapin")
        if all_l2.empty: st.info("Aucun lapin enregistré.")
        else:
            choix=st.selectbox("Choisir un lapin",all_l2['nom'].tolist())
            row=read_sql("SELECT * FROM lapin WHERE nom=%s", [choix]).iloc[0]
            id_l=int(row['idlapin'])
            b_st=statut_badge(row['statut'])
            b_sx=badge("♂️ Mâle" if row['sexe']=='M' else "♀️ Femelle","bml" if row['sexe']=='M' else "bf")

            col_ph, col_info = st.columns([1,3])
            with col_ph:
                if row.get('photo'):
                    try:
                        img_data = base64.b64decode(row['photo'])
                        st.image(img_data, width=180)
                    except: st.markdown('<div style="width:140px;height:140px;background:#DCFCE7;border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:3rem;">🐇</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div style="width:140px;height:140px;background:#DCFCE7;border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:3rem;">🐇</div>', unsafe_allow_html=True)

            with col_info:
                st.markdown(f"""
                <div class="fiche"><div class="fiche-top">
                    <div class="fiche-av">🐇</div>
                    <div><div class="fiche-nom">{row['nom']}</div>
                    <div class="fiche-id">ID #{row['idlapin']} &nbsp;·&nbsp; {b_st} {b_sx}</div></div>
                </div><div class="fiche-body">
                <div class="fiche-grid">
                    <div class="ff"><label>Race</label><span>{row['race'] or '—'}</span></div>
                    <div class="ff"><label>Couleur</label><span>{row['couleur'] or '—'}</span></div>
                    <div class="ff"><label>Naissance</label><span>{row['datenaissance'] or '—'}</span></div>
                    <div class="ff"><label>Origine</label><span>{row['origine'] or '—'}</span></div>
                    <div class="ff"><label>Père (ID)</label><span>{row['idpere'] or '—'}</span></div>
                    <div class="ff"><label>Mère (ID)</label><span>{row['idmere'] or '—'}</span></div>
                </div>
                <div class="ff"><label>Remarques</label><span>{row['remarques'] or '—'}</span></div>
                </div></div>""", unsafe_allow_html=True)

            cg, cs = st.columns(2)
            with cg:
                df_pw=read_sql("SELECT datepesee,poids FROM pesee WHERE idlapin=%s ORDER BY datepesee", [id_l])
                if not df_pw.empty:
                    f=px.line(df_pw,x='datepesee',y='poids',title=f"Croissance — {choix}",markers=True,color_discrete_sequence=['#16A34A'])
                    f.update_layout(xaxis_title='Date',yaxis_title='Poids (g)')
                    pchart(f)
                else: st.info("Aucune pesée.")
            with cs:
                df_sa=read_sql("SELECT datetraitement,typetraitement,produit FROM sante WHERE idlapin=%s ORDER BY datetraitement DESC", [id_l])
                if not df_sa.empty:
                    st.markdown("**💊 Santé**")
                    st.dataframe(df_sa,use_container_width=True)
                else: st.info("Aucun traitement.")

            # Exports fiche
            st.markdown("<hr>", unsafe_allow_html=True)
            ce1, ce2, ce3 = st.columns(3)
            with ce1:
                pdf_b=export_pdf_fiche(row, df_pw if 'df_pw' in dir() else pd.DataFrame(), df_sa if 'df_sa' in dir() else pd.DataFrame())
                st.download_button(f"📄 Fiche PDF",pdf_b,f"fiche_{choix}.pdf","application/pdf")
            with ce2:
                df_exp=pd.DataFrame([{"Nom":row['nom'],"Sexe":row['sexe'],"Race":row['race'],"Statut":row['statut']}])
                xls=export_excel({"Fiche":df_exp})
                st.download_button("📊 Fiche Excel",xls,f"fiche_{choix}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with ce3:
                if role_ok("employe"):
                    new_st=st.selectbox("Statut",["vivant","vendu","mort"],index=["vivant","vendu","mort"].index(row['statut']))
                    if st.button("💾 Mettre à jour"):
                        run_query("UPDATE lapin SET statut=%s WHERE idlapin=%s",(new_st,id_l))
                        # committed via run_query
                        log_action("MAJ_STATUT",f"{choix} → {new_st}")
                        st.success("✅ Mis à jour !"); st.rerun()

    with tabs[3]:
        st.markdown("### 🌳 Arbre généalogique")
        all_l3=read_sql("SELECT idlapin,nom FROM lapin")
        if all_l3.empty: st.info("Aucun lapin.")
        else:
            lapin_sel=st.selectbox("Lapin central",all_l3['nom'].tolist(),key="gen_sel")
            row_g=read_sql("SELECT * FROM lapin WHERE nom=%s", [lapin_sel]).iloc[0]
            pere_nom=run_query("SELECT nom FROM lapin WHERE idlapin=%s",(row_g['idpere'],)).fetchone() if row_g['idpere'] else None
            mere_nom=run_query("SELECT nom FROM lapin WHERE idlapin=%s",(row_g['idmere'],)).fetchone() if row_g['idmere'] else None
            enfants=read_sql("SELECT nom,sexe FROM lapin WHERE idpere=%s OR idmere=%s", [int(row_g['idlapin']),int(row_g['idlapin'])])

            # Visualisation simple avec Plotly
            nodes_x, nodes_y, labels, colors_n = [], [], [], []
            annotations = []

            def add_node(x, y, label, color):
                nodes_x.append(x); nodes_y.append(y)
                labels.append(label); colors_n.append(color)

            add_node(0, 0, f"🐇 {lapin_sel}\n(Principal)", "#16A34A")
            lines_x, lines_y = [], []
            if pere_nom:
                add_node(-1.5, 1, f"♂️ {pere_nom[0]}\n(Père)", "#1E40AF")
                lines_x += [-1.5, 0, None]; lines_y += [1, 0, None]
            if mere_nom:
                add_node(1.5, 1, f"♀️ {mere_nom[0]}\n(Mère)", "#9D174D")
                lines_x += [1.5, 0, None]; lines_y += [1, 0, None]
            for i, (_, e) in enumerate(enfants.iterrows()):
                cx = (i - len(enfants)/2) * 1.2
                add_node(cx, -1, f"{'♂️' if e['sexe']=='M' else '♀️'} {e['nom']}", "#64748B")
                lines_x += [cx, 0, None]; lines_y += [-1, 0, None]

            fig_tree = go.Figure()
            if lines_x:
                fig_tree.add_trace(go.Scatter(x=lines_x, y=lines_y, mode='lines',
                                               line=dict(color='#CBD5E1', width=2), showlegend=False, hoverinfo='skip'))
            fig_tree.add_trace(go.Scatter(
                x=nodes_x, y=nodes_y, mode='markers+text',
                marker=dict(size=40, color=colors_n, line=dict(color='white', width=2)),
                text=labels, textposition='middle center',
                textfont=dict(size=9, color='white', family='Sora'),
                showlegend=False,
                hoverinfo='text'
            ))
            fig_tree.update_layout(
                title=f"Arbre généalogique — {lapin_sel}",
                xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                height=400, **PLOTLY
            )
            pchart(fig_tree)
            if not enfants.empty:
                st.markdown(f"**Descendants directs :** {', '.join(enfants['nom'].tolist())}")

    # connection pooled

# ═══════════════════════════════════════════════════════
# REPRODUCTION
# ═══════════════════════════════════════════════════════
def page_reproduction():
    ph("❤️","Reproduction","Suivi des accouplements et gestations")
    conn=get_conn()
    males=read_sql("SELECT idlapin,nom FROM lapin WHERE sexe='M' AND statut='vivant'")
    femelles=read_sql("SELECT idlapin,nom FROM lapin WHERE sexe='F' AND statut='vivant'")

    with st.expander("➕ Enregistrer un accouplement",expanded=True):
        with st.form("frep"):
            if males.empty or femelles.empty:
                st.warning("Il faut au moins un mâle et une femelle vivants.")
                st.form_submit_button("Enregistrer",disabled=True)
            else:
                c1,c2=st.columns(2)
                ms=c1.selectbox("♂️ Mâle",males['nom'].tolist())
                fs=c2.selectbox("♀️ Femelle",femelles['nom'].tolist())
                c3,c4,c5=st.columns(3)
                da=c3.date_input("Date accouplement",value=date.today())
                dc=c4.date_input("Contrôle (J+14)",value=date.today()+timedelta(days=14))
                dm=c5.date_input("Mise bas prévue (J+31)",value=date.today()+timedelta(days=31))
                notes=st.text_area("Notes")
                if st.form_submit_button("💾 Enregistrer",use_container_width=True):
                    im=int(males[males.nom==ms]['idlapin'].values[0])
                    if_=int(femelles[femelles.nom==fs]['idlapin'].values[0])
                    run_query("INSERT INTO reproduction (idmale,idfemelle,dateaccouplement,datecontrole,datemisebas,notes) VALUES (%s,%s,%s,%s,%s,%s)",
                                 (im,if_,str(da),str(dc),str(dm),notes))
                    # committed via run_query
                    log_action("ACCOUPLEMENT",f"{ms} × {fs}")
                    st.success("✅ Accouplement enregistré !")

    # Confirmer gestation
    df_rep=read_sql("""SELECT r.idrepro, lm.nom||' × '||lf.nom AS Couple,
        r.dateaccouplement, r.datecontrole, r.datemisebas, r.gestationconfirmee
        FROM reproduction r JOIN lapin lm ON r.idmale=lm.idlapin JOIN lapin lf ON r.idfemelle=lf.idlapin
        ORDER BY r.dateaccouplement DESC""")

    st.subheader("Historique — Confirmer gestation")
    if not df_rep.empty:
        for _,r in df_rep.iterrows():
            cc1,cc2,cc3=st.columns([3,1,1])
            cc1.markdown(f"**{r['couple']}** · {r['dateaccouplement']}")
            cc2.markdown(badge("✅ Confirmée","bv") if r['gestationconfirmee'] else badge("⏳ En attente","bw"), unsafe_allow_html=True)
            if not r['gestationconfirmee']:
                if cc3.button("Confirmer", key=f"conf_{r['idrepro']}"):
                    run_query("UPDATE reproduction SET gestationconfirmee=1 WHERE idrepro=%s",(int(r['idrepro']),))
                    # committed via run_query
                    st.rerun()

    xls=export_excel({"Reproductions":df_rep})
    st.download_button("📊 Exporter Excel",xls,"reproductions.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # connection pooled

# ═══════════════════════════════════════════════════════
# PORTÉES
# ═══════════════════════════════════════════════════════
def page_portees():
    ph("👶","Portées & Naissances","Enregistrement des mises bas")
    conn=get_conn()
    repros=read_sql("SELECT r.idrepro, lm.nom||' × '||lf.nom AS Couple FROM reproduction r JOIN lapin lm ON r.idmale=lm.idlapin JOIN lapin lf ON r.idfemelle=lf.idlapin")
    with st.expander("➕ Enregistrer une portée",expanded=True):
        with st.form("fpor"):
            if repros.empty:
                st.warning("Aucune reproduction enregistrée.")
                st.form_submit_button("Enregistrer",disabled=True)
            else:
                rs=st.selectbox("Reproduction",repros['couple'].tolist())
                c1,c2,c3=st.columns(3)
                dn=c1.date_input("Date naissance",value=date.today())
                nv=c2.number_input("Nés vivants",min_value=0,step=1)
                nm=c3.number_input("Mort-nés",min_value=0,step=1)
                notes=st.text_area("Notes")
                if st.form_submit_button("💾 Enregistrer",use_container_width=True):
                    ir=int(repros[repros.Couple==rs]['idrepro'].values[0])
                    run_query("INSERT INTO portee (idrepro,datenaissance,nbvivant,nbmort,notes) VALUES (%s,%s,%s,%s,%s)",(ir,str(dn),nv,nm,notes))
                    # committed via run_query
                    log_action("PORTEE",f"{rs}: {nv} vivants, {nm} morts")
                    st.success(f"✅ {nv} vivants, {nm} mort-nés enregistrés.")
    df=read_sql("SELECT lm.nom||' × '||lf.nom AS Parents, p.datenaissance AS Date, p.nbvivant AS Vivants, p.nbmort AS MortNés, (p.nbvivant+p.nbmort) AS Total FROM portee p JOIN reproduction r ON p.idrepro=r.idrepro JOIN lapin lm ON r.idmale=lm.idlapin JOIN lapin lf ON r.idfemelle=lf.idlapin ORDER BY p.datenaissance DESC",conn)
    st.subheader("Historique des portées")
    st.dataframe(df,use_container_width=True)
    xls=export_excel({"Portées":df})
    st.download_button("📊 Exporter Excel",xls,"portees.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # connection pooled

# ═══════════════════════════════════════════════════════
# PESÉES
# ═══════════════════════════════════════════════════════
def page_pesees():
    ph("⚖️","Pesées & Croissance","Suivi pondéral du cheptel")
    conn=get_conn()
    lapins=read_sql("SELECT idlapin,nom FROM lapin WHERE statut='vivant'")
    with st.expander("➕ Ajouter une pesée",expanded=True):
        with st.form("fpes"):
            if lapins.empty:
                st.warning("Aucun lapin vivant."); st.form_submit_button("Enregistrer",disabled=True)
            else:
                c1,c2,c3=st.columns(3)
                ls=c1.selectbox("Lapin",lapins['nom'].tolist())
                dp=c2.date_input("Date",value=date.today())
                po=c3.number_input("Poids (g)",min_value=0.0,step=10.0)
                if st.form_submit_button("💾 Enregistrer",use_container_width=True):
                    il=int(lapins[lapins.nom==ls]['idlapin'].values[0])
                    run_query("INSERT INTO pesee (idlapin,datepesee,poids) VALUES (%s,%s,%s)",(il,str(dp),po))
                    # committed via run_query
                    log_action("PESEE",f"{ls}: {po}g")
                    st.success(f"✅ {po}g pour {ls}")
    df=read_sql("SELECT l.nom AS Lapin, p.datepesee AS Date, p.poids AS poids_g FROM pesee p JOIN lapin l ON p.idlapin=l.idlapin ORDER BY p.datepesee")
    if not df.empty:
        f=px.line(df,x='Date',y='poids_g',color='Lapin',markers=True,color_discrete_sequence=GREENS,title="Courbes de croissance")
        f.update_layout(xaxis_title='Date',yaxis_title='Poids (g)')
        pchart(f)
    st.subheader("Historique")
    st.dataframe(df,use_container_width=True)
    xls=export_excel({"Pesées":df})
    st.download_button("📊 Exporter Excel",xls,"pesees.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # connection pooled

# ═══════════════════════════════════════════════════════
# SANTÉ
# ═══════════════════════════════════════════════════════
def page_sante():
    ph("💊","Santé & Traitements","Carnet de santé de votre cheptel")
    conn=get_conn()
    lapins=read_sql("SELECT idlapin,nom FROM lapin WHERE statut='vivant'")
    with st.expander("➕ Ajouter un traitement",expanded=True):
        with st.form("fsan"):
            if lapins.empty:
                st.warning("Aucun lapin vivant."); st.form_submit_button("Enregistrer",disabled=True)
            else:
                c1,c2=st.columns(2)
                ls=c1.selectbox("Lapin",lapins['nom'].tolist())
                dt=c2.date_input("Date",value=date.today())
                c3,c4=st.columns(2)
                tt=c3.selectbox("Type",["Vaccin","Antiparasite","Antibiotique","Vitamine","Autre"])
                pr=c4.text_input("Produit")
                c5,c6=st.columns(2)
                rappel=c5.date_input("Date de rappel (optionnel)",value=None)
                rm=st.text_area("Remarques")
                if st.form_submit_button("💾 Enregistrer",use_container_width=True):
                    il=int(lapins[lapins.nom==ls]['idlapin'].values[0])
                    run_query("INSERT INTO sante (idlapin,datetraitement,typetraitement,produit,remarque,daterappel) VALUES (%s,%s,%s,%s,%s,%s)",
                                 (il,str(dt),tt,pr,rm,str(rappel) if rappel else None))
                    # committed via run_query
                    log_action("SANTE",f"{ls}: {tt} — {pr}")
                    st.success("✅ Traitement enregistré !")
    df=read_sql("SELECT l.nom AS Lapin, s.datetraitement AS Date, s.typetraitement AS Type, s.produit AS Produit, s.daterappel AS Rappel, s.remarque AS Remarques FROM sante s JOIN lapin l ON s.idlapin=l.idlapin ORDER BY s.datetraitement DESC")
    st.subheader("Historique des traitements")
    st.dataframe(df,use_container_width=True)
    xls=export_excel({"Santé":df})
    st.download_button("📊 Exporter Excel",xls,"sante.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # connection pooled

# ═══════════════════════════════════════════════════════
# VENTES
# ═══════════════════════════════════════════════════════
def page_ventes():
    ph("💰","Ventes","Gestion des transactions et revenus")
    conn=get_conn()
    lapins=read_sql("SELECT idlapin,nom FROM lapin WHERE statut='vivant'")
    with st.expander("➕ Enregistrer une vente",expanded=True):
        with st.form("fven"):
            if lapins.empty:
                st.warning("Aucun lapin disponible."); st.form_submit_button("Enregistrer",disabled=True)
            else:
                c1,c2=st.columns(2)
                ls=c1.selectbox("Lapin",lapins['nom'].tolist())
                dv=c2.date_input("Date",value=date.today())
                c3,c4=st.columns(2)
                px_=c3.number_input("Prix (FCFA)",min_value=0.0,step=100.0)
                cl=c4.text_input("Client")
                if st.form_submit_button("💾 Enregistrer",use_container_width=True):
                    il=int(lapins[lapins.nom==ls]['idlapin'].values[0])
                    run_query("INSERT INTO vente (idlapin,datevente,prix,client) VALUES (%s,%s,%s,%s)",(il,str(dv),px_,cl))
                    run_query("UPDATE lapin SET statut='vendu' WHERE idlapin=%s",(il,))
                    # committed via run_query
                    log_action("VENTE",f"{ls}: {px_:,.0f} FCFA à {cl}")
                    st.success(f"✅ {ls} vendu — {px_:,.0f} FCFA")
    df=read_sql("SELECT l.nom AS Lapin, v.datevente AS Date, v.prix AS prix_fcfa, v.client AS Client FROM vente v JOIN lapin l ON v.idlapin=l.idlapin ORDER BY v.datevente DESC")
    if not df.empty:
        c1,c2=st.columns(2)
        c1.metric("💵 Total revenus",f"{df['prix_fcfa'].sum():,.0f} FCFA")
        c2.metric("📦 Ventes totales",len(df))
    st.subheader("Historique")
    st.dataframe(df,use_container_width=True)
    xls=export_excel({"Ventes":df})
    st.download_button("📊 Exporter Excel",xls,"ventes.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # connection pooled

# ═══════════════════════════════════════════════════════
# STOCKS
# ═══════════════════════════════════════════════════════
def page_stocks():
    ph("📦","Stocks & Coûts","Gestion des aliments, médicaments et dépenses")
    conn=get_conn()
    tabs=st.tabs(["📋 Inventaire","➕ Ajouter","📊 Coûts"])
    with tabs[0]:
        df=read_sql("SELECT * FROM stock ORDER BY dateachat DESC")
        if not df.empty:
            total_val=sum(r['quantite']*r['prixunitaire'] for _,r in df.iterrows())
            st.metric("💰 Valeur totale du stock",f"{total_val:,.0f} FCFA")
        st.dataframe(df,use_container_width=True)
        if not df.empty:
            xls=export_excel({"Stock":df})
            st.download_button("📊 Exporter Excel",xls,"stock.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with tabs[1]:
        with st.form("fstk"):
            c1,c2=st.columns(2)
            tp=c1.selectbox("Type",["Aliment","Médicament","Matériel","Autre"])
            nm=c2.text_input("Nom du produit")
            c3,c4,c5=st.columns(3)
            qt=c3.number_input("Quantité",min_value=0.0,step=0.1)
            ut=c4.selectbox("Unité",["kg","g","L","ml","unité","sac"])
            pu=c5.number_input("Prix unitaire (FCFA)",min_value=0.0,step=10.0)
            da=st.date_input("Date d'achat",value=date.today())
            nt=st.text_area("Notes")
            if st.form_submit_button("💾 Enregistrer",use_container_width=True):
                run_query("INSERT INTO stock (type,nom,quantite,unite,prixunitaire,dateachat,notes) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                             (tp,nm,qt,ut,pu,str(da),nt))
                # committed via run_query
                log_action("STOCK",f"Ajout {nm}: {qt}{ut}")
                st.success(f"✅ {nm} ajouté au stock !")
    with tabs[2]:
        df2=read_sql("SELECT type, SUM(quantite*prixunitaire) as total FROM stock GROUP BY type",conn)
        if not df2.empty:
            f=px.pie(df2,values='total',names='type',title="Répartition des coûts",
                     color_discrete_sequence=GREENS,hole=.35)
            pchart(f)
    # connection pooled

# ═══════════════════════════════════════════════════════
# CALENDRIER
# ═══════════════════════════════════════════════════════
def page_calendrier():
    ph("📅","Calendrier","Agenda des événements à venir")
    conn=get_conn()
    today=date.today()

    evenements=[]
    # Contrôles gestation
    ctrl=read_sql("SELECT r.datecontrole, lm.nom AS male, lf.nom AS femelle FROM reproduction r JOIN lapin lm ON r.idmale=lm.idlapin JOIN lapin lf ON r.idfemelle=lf.idlapin WHERE r.gestationconfirmee=0 AND r.datecontrole IS NOT NULL")
    for _,r in ctrl.iterrows():
        evenements.append({"Date":str(r['datecontrole']),"Type":"🔬 Contrôle gestation","Détail":f"{r['male']} × {r['femelle']}","Priorité":"Moyen"})

    # Mises bas prévues
    mb=read_sql("SELECT r.datemisebas, lm.nom AS male, lf.nom AS femelle FROM reproduction r JOIN lapin lm ON r.idmale=lm.idlapin JOIN lapin lf ON r.idfemelle=lf.idlapin WHERE r.gestationconfirmee=1 AND r.datemisebas IS NOT NULL")
    for _,r in mb.iterrows():
        evenements.append({"Date":str(r['datemisebas']),"Type":"🍼 Mise bas","Détail":f"{r['male']} × {r['femelle']}","Priorité":"Haute"})

    # Rappels vaccin
    rap=read_sql("SELECT l.nom, s.typetraitement, s.daterappel FROM sante s JOIN lapin l ON s.idlapin=l.idlapin WHERE s.daterappel IS NOT NULL")
    for _,r in rap.iterrows():
        evenements.append({"Date":str(r['daterappel']),"Type":f"💊 {r['typetraitement']}","Détail":r['nom'],"Priorité":"Moyen"})

    if evenements:
        df_cal=pd.DataFrame(evenements).sort_values("Date")
        df_cal=df_cal[df_cal['Date']>=str(today)]
        if not df_cal.empty:
            st.markdown(f"**{len(df_cal)} événement(s) à venir**")
            for _,e in df_cal.iterrows():
                try:
                    d=datetime.strptime(e['Date'],'%Y-%m-%d').date()
                    diff=(d-today).days
                    urgence="alert-danger" if diff<=3 else "alert-warn" if diff<=7 else "alert-info"
                    icon="🔴" if diff<=3 else "🟡" if diff<=7 else "🔵"
                    st.markdown(f'<div class="alert-card {urgence}"><div class="alert-icon">{icon}</div><div><div class="alert-title">{e["Type"]} — {e["Détail"]}</div><div class="alert-sub">📅 {d.strftime("%d/%m/%Y")} · Dans {diff} jour(s)</div></div></div>',unsafe_allow_html=True)
                except: pass
        else:
            st.success("✅ Aucun événement à venir.")
    else:
        st.info("Aucun événement planifié pour l'instant.")
    # connection pooled

# ═══════════════════════════════════════════════════════
# ASSISTANT IA
# ═══════════════════════════════════════════════════════
def page_assistant():
    ph("🤖","Assistant IA","Posez vos questions à l'IA sur votre élevage")

    if "chat_history" not in st.session_state:
        st.session_state.chat_history=[]

    ctx=get_elevage_context()
    system=f"""Tu es un assistant expert en cuniculture (élevage de lapins) pour l'Élevage Ouakoubo en Côte d'Ivoire.
Tu as accès aux données réelles de l'élevage :
- Total lapins vivants : {ctx['total_vivants']} ({ctx['males']} mâles, {ctx['femelles']} femelles)
- Portées totales : {ctx['total_portees']}
- Revenus totaux : {ctx['revenus_total']:,.0f} FCFA
- Races élevées : {', '.join(ctx['races']) if ctx['races'] else 'Non renseignées'}
- Meilleure reproductrice : {ctx['meilleure_femelle']['nom'] if ctx['meilleure_femelle'] else '—'} ({ctx['meilleure_femelle']['vivants'] if ctx['meilleure_femelle'] else 0} vivants)

Réponds en français, de manière concise et pratique. Tu peux donner des conseils vétérinaires généraux, 
analyser les données de l'élevage, suggérer des améliorations. 
Sois chaleureux et professionnel. Si on te demande une analyse, base-toi sur les données ci-dessus."""

    # Affichage historique
    for msg in st.session_state.chat_history:
        if msg['role']=='user':
            st.markdown(f'<div class="chat-label" style="text-align:right;color:#64748B;">Vous</div><div class="chat-user">{msg["content"]}</div>',unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="chat-label" style="color:#16A34A;">🤖 Assistant IA</div><div class="chat-ai">{msg["content"]}</div>',unsafe_allow_html=True)

    # Suggestions rapides
    if not st.session_state.chat_history:
        st.markdown("**💡 Questions suggérées :**")
        sugg=["Quelle est la meilleure femelle reproductrice ?","Donne-moi un résumé de mon élevage","Conseils pour améliorer le taux de survie des lapereaux","Quel est mon chiffre d'affaires ?","Comment détecter une maladie chez un lapin ?"]
        cols=st.columns(len(sugg))
        for i,s in enumerate(sugg):
            if cols[i].button(s, key=f"sugg_{i}", use_container_width=True):
                st.session_state.chat_history.append({"role":"user","content":s})
                with st.spinner("L'IA réfléchit..."):
                    rep=call_claude(st.session_state.chat_history, system)
                st.session_state.chat_history.append({"role":"assistant","content":rep})
                st.rerun()

    # Input
    with st.form("chat_form",clear_on_submit=True):
        c1,c2=st.columns([5,1])
        user_input=c1.text_input("Votre question...",label_visibility="collapsed",placeholder="Ex: Quel lapin n'a pas été pesé depuis 2 semaines ?")
        send=c2.form_submit_button("Envoyer 📤",use_container_width=True)
    if send and user_input.strip():
        st.session_state.chat_history.append({"role":"user","content":user_input})
        with st.spinner("L'IA réfléchit..."):
            rep=call_claude(st.session_state.chat_history[-6:], system)
        st.session_state.chat_history.append({"role":"assistant","content":rep})
        st.rerun()

    if st.session_state.chat_history:
        if st.button("🗑️ Effacer la conversation"):
            st.session_state.chat_history=[]
            st.rerun()

# ═══════════════════════════════════════════════════════
# MULTI-UTILISATEURS
# ═══════════════════════════════════════════════════════
def page_utilisateurs():
    ph("👥","Gestion des Utilisateurs","Comptes et permissions")
    if not role_ok("admin"): return
    conn=get_conn()

    tabs=st.tabs(["👤 Liste","➕ Ajouter"])
    with tabs[0]:
        df=read_sql("SELECT id, login, role, datecreation FROM utilisateur")
        st.dataframe(df,use_container_width=True)

    with tabs[1]:
        with st.form("fusr"):
            c1,c2=st.columns(2)
            new_login=c1.text_input("Identifiant")
            new_role=c2.selectbox("Rôle",["admin","employe","consultant"])
            new_pwd=st.text_input("Mot de passe",type="password")
            if st.form_submit_button("➕ Créer l'utilisateur",use_container_width=True):
                if not new_login or not new_pwd:
                    st.error("Identifiant et mot de passe obligatoires.")
                elif len(new_pwd)<6:
                    st.error("Mot de passe trop court (min 6 caractères).")
                else:
                    h=hashlib.sha256(new_pwd.encode()).hexdigest()
                    try:
                        run_query("INSERT INTO utilisateur (login,password,role) VALUES (%s,%s,%s)",(new_login,h,new_role))
                        # committed via run_query
                        log_action("CREATION_USER",f"Utilisateur {new_login} ({new_role}) créé")
                        st.success(f"✅ Utilisateur **{new_login}** créé avec le rôle **{new_role}** !")
                    except: st.error("❌ Cet identifiant existe déjà.")

    st.markdown("<hr>",unsafe_allow_html=True)
    st.markdown("""
    **📖 Rôles disponibles**
    - **admin** : Accès complet (lecture, écriture, suppression, gestion utilisateurs)
    - **employe** : Saisie des données (ajout, modification) — pas de suppression ni gestion users
    - **consultant** : Lecture seule (tableau de bord, rapports) — idéal pour vétérinaire / investisseur
    """)
    # connection pooled

# ═══════════════════════════════════════════════════════
# JOURNAL
# ═══════════════════════════════════════════════════════
def page_journal():
    ph("📋","Journal d'activité","Traçabilité de toutes les actions")
    if not role_ok("admin"): return
    conn=get_conn()
    df=read_sql("SELECT dateaction, utilisateur, action, detail FROM journal ORDER BY dateaction DESC LIMIT 200")
    st.dataframe(df,use_container_width=True,height=500)
    xls=export_excel({"Journal":df})
    st.download_button("📊 Exporter Journal",xls,"journal.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # connection pooled

# ═══════════════════════════════════════════════════════
# PARAMÈTRES
# ═══════════════════════════════════════════════════════
def page_parametres():
    ph("⚙️","Paramètres","Configuration du compte et du système")
    conn=get_conn()
    st.subheader("🔑 Changer le mot de passe")
    with st.form("fpwd"):
        a=st.text_input("Ancien mot de passe",type="password")
        n1=st.text_input("Nouveau mot de passe",type="password")
        n2=st.text_input("Confirmer",type="password")
        if st.form_submit_button("🔒 Mettre à jour",use_container_width=True):
            ah=hashlib.sha256(a.encode()).hexdigest()
            user=run_query("SELECT * FROM utilisateur WHERE login=%s AND password=%s",(st.session_state['username'],ah)).fetchone()
            if not user: st.error("❌ Ancien mot de passe incorrect.")
            elif n1!=n2: st.error("❌ Mots de passe différents.")
            elif len(n1)<6: st.error("❌ Minimum 6 caractères.")
            else:
                nh=hashlib.sha256(n1.encode()).hexdigest()
                run_query("UPDATE utilisateur SET password=%s WHERE login=%s",(nh,st.session_state['username']))
                # committed via run_query
                log_action("MAJ_PASSWORD","Mot de passe modifié")
                st.success("✅ Mot de passe mis à jour !")
    st.markdown("<hr>",unsafe_allow_html=True)

    st.info("💾 Sauvegarde : Vos données sont sécurisées sur Supabase (cloud).")

    st.markdown("<hr>",unsafe_allow_html=True)
    st.markdown("""
**ℹ️ À propos**

🐇 **Élevage Ouakoubo** — Système de gestion cuniculicole
`Version 4.0` · Streamlit + SQLite + Plotly + Claude AI

✅ Export Excel & PDF · 🔔 Alertes automatiques · 🌳 Généalogie
👥 Multi-utilisateurs · 🤖 Assistant IA · 📦 Gestion stocks
📅 Calendrier · 📋 Journal d'activité · 📸 Photos lapins

Développé avec ❤️ · Abidjan, Côte d'Ivoire
    """)
    # connection pooled

# ═══════════════════════════════════════════════════════
# NAVIGATION — Accordéon avec barres cliquables
# ═══════════════════════════════════════════════════════

# Définition des sections et pages (sans Assistant IA)
NAV_STRUCTURE = {
    "Principal": [
        ("dashboard",   "📊", "Tableau de bord"),
        ("alertes",     "🔔", "Alertes"),
        ("calendrier",  "📅", "Calendrier"),
    ],
    "Élevage": [
        ("lapins",       "🐇", "Lapins"),
        ("reproduction", "❤️", "Reproduction"),
        ("portees",      "👶", "Portées"),
        ("pesees",       "⚖️", "Pesées"),
        ("sante",        "💊", "Santé"),
        ("ventes",       "💰", "Ventes"),
        ("stocks",       "📦", "Stocks"),
    ],
    "Administration": [
        ("utilisateurs", "👥", "Utilisateurs"),
        ("journal",      "📋", "Journal"),
        ("parametres",   "⚙️", "Paramètres"),
    ],
}

PAGE_MAP = {
    "dashboard":    page_dashboard,
    "alertes":      page_alertes,
    "calendrier":   page_calendrier,
    "lapins":       page_lapins,
    "reproduction": page_reproduction,
    "portees":      page_portees,
    "pesees":       page_pesees,
    "sante":        page_sante,
    "ventes":       page_ventes,
    "stocks":       page_stocks,
    "utilisateurs": page_utilisateurs,
    "journal":      page_journal,
    "parametres":   page_parametres,
}

# Section par défaut ouverte selon la page active
def get_open_section(page_id):
    for section, items in NAV_STRUCTURE.items():
        if any(p[0] == page_id for p in items):
            return section
    return "Principal"

def render_sidebar(alertes_count):
    """Rendu de la sidebar avec accordéon en HTML + boutons Streamlit invisibles."""

    active_page = st.session_state.get("current_page", "dashboard")
    open_section = st.session_state.get("open_section", get_open_section(active_page))
    role_label   = {"admin":"👑 Admin","employe":"🔧 Employé","consultant":"👁️ Consultant"}.get(
                    st.session_state.get("role",""), "")

    # ── Logo ──
    st.markdown(f"""
    <div class="sb-logo">
      <div class="sb-logo-inner">
        <div class="sb-logo-icon">🐇</div>
        <div>
          <div class="sb-logo-name">Élevage Ouakoubo</div>
          <div class="sb-logo-sub">Gestion cuniculicole v5.0</div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Sections accordéon ──
    for section, items in NAV_STRUCTURE.items():
        is_open = (open_section == section)
        arrow   = "▶"

        # En-tête de section cliquable (bouton Streamlit invisible)
        col_hdr, col_btn = st.columns([5, 1])
        with col_hdr:
            open_cls = "open" if is_open else ""
            st.markdown(f"""
            <div class="sb-acc-header {open_cls}">
              <span>{section.upper()}</span>
              <span class="sb-acc-arrow">{arrow}</span>
            </div>
            """, unsafe_allow_html=True)
        with col_btn:
            # Bouton transparent pour toggle
            st.markdown('<div style="margin-top:-40px;opacity:0;">', unsafe_allow_html=True)
            if st.button("_", key=f"sec_{section}", use_container_width=True):
                st.session_state["open_section"] = section if not is_open else ""
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        # Items de la section (visibles seulement si ouverte)
        if is_open:
            for page_id, icon, label in items:
                is_active = (active_page == page_id)
                active_cls = "active" if is_active else ""

                # Badge alerte sur "alertes"
                badge_html = ""
                if page_id == "alertes" and alertes_count > 0:
                    badge_html = f'<span class="sb-nav-badge">{alertes_count}</span>'

                st.markdown(f"""
                <div class="sb-nav-item {active_cls}" id="nav_{page_id}">
                  <span class="sb-nav-icon">{icon}</span>
                  <span>{label}</span>
                  {badge_html}
                </div>
                """, unsafe_allow_html=True)

                # Bouton invisible qui couvre la barre
                st.markdown('<div style="margin-top:-38px;opacity:0;">', unsafe_allow_html=True)
                if st.button(label, key=f"nav_{page_id}", use_container_width=True):
                    st.session_state["current_page"] = page_id
                    st.session_state["open_section"] = section
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

    # ── Utilisateur + Déconnexion ──
    st.markdown(f"""
    <div class="sb-user">
      <div class="sb-user-av">👤</div>
      <div>
        <div class="sb-user-name">{st.session_state.get('username','')}</div>
        <div class="sb-user-role">{role_label}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    if st.button("🚪 Déconnexion", key="btn_logout", use_container_width=True):
        log_action("DECONNEXION")
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()


# ═══════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════
st.set_page_config(
    page_title="Élevage Ouakoubo",
    page_icon="🐇",
    layout="wide",
    initial_sidebar_state="expanded"
)
init_db()
inject_css()

# Init session state
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
if "current_page" not in st.session_state:
    st.session_state["current_page"] = "dashboard"
if "open_section" not in st.session_state:
    st.session_state["open_section"] = "Principal"

if not st.session_state["logged_in"]:
    login_page()
else:
    # Calculer alertes une seule fois
    alertes_count = len(get_alertes())

    with st.sidebar:
        render_sidebar(alertes_count)

    # Afficher la page active
    current = st.session_state.get("current_page", "dashboard")
    page_fn = PAGE_MAP.get(current, page_dashboard)
    page_fn()
